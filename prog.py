import openpyxl

# Open the workbook
wb = openpyxl.load_workbook('student_data.xlsx')

# Get the active sheet
ws = wb.active

# Add some data to the sheet
ws['A1'] = 'Name'
ws['B1'] = 'Roll Number'
ws['C1'] = 'Section'
ws['A2'] = 'Pushpa'
ws['B2'] = '4389'
ws['C2'] = 'CAI'
ws['A3'] = 'Anjali'
ws['B3'] = '4364'
ws['C3'] = 'CAI'
ws['A4'] = 'Yogitha'
ws['B4'] = '4392'
ws['C4'] = 'CAI'+
ws['A5'] = 'Siri'
ws['B5'] = '45E9'
ws['C5'] = 'AID'
ws['A6'] = 'Appu'
ws['B6'] = '45E9'
ws['C6'] = 'AID'
# Save the workbook
wb.save('student_data.xlsx')